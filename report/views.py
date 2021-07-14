from datetime import datetime
from django.shortcuts import render
from django.http import HttpRequest
from django.contrib.auth.decorators import login_required
from report.models import *
from django.views.generic import TemplateView
from app.views import home
from django.urls import *
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.http import JsonResponse
from django.http import HttpResponse
import numpy as np
import math as mt
from app.models import *
import json
from django.db.models import Sum
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.table import Table as TableExcel
from openpyxl.worksheet.table import TableStyleInfo
import re
from reportlab.pdfgen import canvas
from django.conf import settings
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Paragraph
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, portrait, A4
from reportlab.lib import utils 
import io
from django.core.files.base import ContentFile
from report.models import Image as ImgReport
from report.forms import UploadFileForm, HistoricalForm
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER
from reportlab.pdfbase.pdfmetrics import stringWidth
from django.core.paginator import Paginator
import os
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import redirect

# Create your views here.

#def group_required(*group_names):
#    """Requires user membership in at least one of the groups passed in."""
#    def in_groups(u):
#        if u.is_authenticated():
#            if bool(u.groups.filter(name__in=group_names)) | u.is_superuser:
#                return True
#        return False

#    return user_passes_test(in_groups, login_url='403')

###FORMULARIO DE GENERATE REPORT
@login_required(login_url="login")
def createreport(request):

    #Objetos a mostrar
    apis = API.objects.all()
    contracts = Contract.objects.all()
    specialties = Specialty.objects.all()
    references = Reference.objects.all()
    activities = Activity.objects.all()
    subactivities = SubActivity.objects.all()
    equipments = Equipment.objects.all()
    nonconformities = NonConformity.objects.all()
    followings = Following.objects.all()
    preconditions = Precondition.objects.all()
    measures = Measure.objects.all()

    assert isinstance(request, HttpRequest)
    return render(
        request,
        'form-report/generatereport.html',
        {
            'title':'Registro de Informe Diario',
            'year': datetime.now().year,
            'apis': apis,
            'contracts': contracts,
            'specialties': specialties,
            'references': references,
            'activities': activities,
            'subactivities': subactivities,
            'equipments': equipments,
            'nonconformities': nonconformities,
            'followings': followings,
            'preconditions': preconditions,
            'measures': measures,
        })

@login_required(login_url="login")
def buscarcontratos(request, *args, **kwargs):
    data = {}

    try:
       if request.method == 'GET':
           Id = request.GET['id']
           action = request.GET['action']
           data = []

           if action == 'search_contract_id':
                contrato = Contract.objects.filter(api=Id)
                api = API.objects.filter(id=Id)
                for i in contrato:
                    data.append({'id': i.id, 
                                 'contract_number': i.contract_number,
                                 'project_name': api[0].project_name})

           if action == 'search_contract_info':
               contrato = Contract.objects.filter(id=Id)
               actividad = Activity.objects.filter(contract=Id)
               subactividad = SubActivity.objects.filter(contract=Id)
               for i in actividad:
                   for s in subactividad:
                       data.append({'id_contract': contrato[0].id,
                                    'enterprise': contrato[0].enterprise,
                                    'start_date_contract': contrato[0].start_date,
                                    'finish_date_contract': contrato[0].finish_date,
                                    'id': i.id,
                                    'activity_name': i.activity_name,
                                    'start_date_activity': i.start_date,
                                    'finish_date_activity': i.finish_date,
                                    'subactivity_name': s.subactivity_name,
                                    'subactivity_id': s.id})

           if ((action == 'search_subactivity') or (action == 'search_subactivity_nop')):
               subactividad = SubActivity.objects.filter(activity=Id)

               for i in subactividad:
                   data.append({'id':i.id,
                                'subactivity_name':i.subactivity_name})

           if ((action == 'search_sub_nop_info') or (action == 'search_sub_p_info')):

                subactividad = SubActivity.objects.filter(id=Id)
                medida = Measure.objects.filter(id=subactividad[0].measureunit.id)
                historical = Historical.objects.filter(subactivity=Id)

                if(len(historical)>0):
                    sum = 0
                
                    for i in historical:
                        sum += i.real_amount

                    dias = subactividad[0].finish_date - subactividad[0].start_date
                    diaspro = dias.days            
                    if diaspro == 0:
                        diaspro = 1

                    data.append({'id': subactividad[0].id,
                                 'measure_id': subactividad[0].measureunit.id,
                                 'measure': subactividad[0].measureunit.measure_name,
                                 'average_amount': subactividad[0].average_amount,
                                 'total': sum,
                                 'ref_day': round(subactividad[0].average_amount/diaspro
                                                , 2)
                                 })

                else:
                    data.append({'id': subactividad[0].id,
                                 'measure_id': subactividad[0].measureunit.id,
                                 'measure': medida[0].measure_name,
                                 'average_amount': subactividad[0].average_amount,
                                 'total': 0,
                                 'ref_day': 0})
                
    except Exception as e:
        data['error'] = str(e)
    return JsonResponse(data, safe=False)

@login_required(login_url="login")
def buscarsubactividades(request, *args, **kwargs):
    data = {}
    try:
       if request.method == 'GET':
           action = request.GET['action']
           act = request.GET['act']
           subact = request.GET['subact']
           data = 0
           #data.append({'id': 1})
           if action == 'search_subactivities':
              subactividad = SubActivity.objects.filter(subactivity_name=str(subact))
              if subactividad != " ":
                  data = 1

    except Exception as e:
        data['error'] = e
    return JsonResponse(data, safe=False)

@csrf_exempt
@login_required(login_url="login")
def savereport(request, *args, **kwargs):

    if request.method == "POST":

        action = request.POST.get('action')
        hd = request.POST.getlist('historico[]')
        rd = request.POST.getlist('referencias[]')
        ed = request.POST.getlist('equipos[]')
        rpd = request.POST.getlist('reporte[]')

        if action == 'save_report':

            data = {}

            historicaldata = json.loads(hd[0])
            referencedata = json.loads(rd[0])
            reportdata = json.loads(rpd[0])
            equipmentdata = json.loads(ed[0])

            rp_id = 0
            for i in reportdata:

                follow = Following.objects.get(pk = int(i['id_seguimiento']))

                reporte = Report(
                    deviation_detected = i['desviacion'],
                    action_plan = i['plandeaccion'],
                    evidence_obs = i['evidencia_obs'],
                    following = follow
                    )

                reporte.save()

                #global id_report

                id_report = reporte.id
                rp_id = reporte.id
                  
            for i in historicaldata:

                subact = SubActivity.objects.get(pk = int(i['id_subactividad']))

                histo = Historical.objects.filter(subactivity = int(i['id_subactividad']))
                diff = 0
                for hs in histo:
                    diff += hs.real_amount

                mea = Measure.objects.get(pk = int(i['id_medida']))
                noncon = NonConformity.objects.get(pk = int(i['id_conformidad']))
                spec = Specialty.objects.get(pk = int(i['id_especialidad']))
                precon = Precondition.objects.get(pk = int(i['id_precondicion']))
                actype = ActivityType.objects.get(pk = int(i['id_actividad_type']))
                activi = Activity.objects.get(pk = int(i['id_actividad']))

                historical = Historical(
                    real_amount = i['cantidad_real'],
                    subactivity_no_program = i['subactivity_no_program'],
                    subactivity = subact,
                    measure = mea,
                    nonconformity = noncon,
                    specialty= spec,
                    precondition= precon,
                    user= request.user,
                    activitytype= actype,
                    report = reporte,
                    no_program_total = i['total_estimado'],
                    no_program_refday = i['referencia_diaria'],
                    no_program_total_acu = i['total_acumulado'],
                    activity = activi,
                    difference = diff
                    )

                historical.save()

            if len(rd) > 0:
                for i in referencedata:
                    ref = Reference.objects.get(pk = int(i['referencia_id']))

                    reference = HistoricalReference(
                        description = i['descripcion'],
                        report = reporte,
                        reference = ref
                        )

                    reference.save()

            if len(ed) > 0:

                for i in equipmentdata:
                    act = Activity.objects.get(pk = int(i['id_actividad']))
                    equi = Equipment.objects.get(pk = int(i['id_equipo']))

                    equipment = EquipmentAmount(
                            equipment_amount = i['cantidad'],
                            direct_endowment = i['dot_directa'],
                            direct_reference = i['dot_referen'],
                            indirect_endowment = i['dot_indirecta'],
                            activity = act,
                            equipment = equi,
                            report = reporte
                        )
                    equipment.save()

            if rp_id > 0:
                    
                data = {
                    'submitted': 1,
                    'id_report': rp_id
                    }

                if reportdata[0]['img_exist'] == 0:

                    create_pdf(rp_id)

            else:
                data = {
                    'submitted': 0
                    }

            return JsonResponse(data)

        
        if len(request.FILES.getlist('image')) > 0:

            image = request.FILES.getlist('image')
            observation = request.POST.getlist('observation')
            subactivity_image = request.POST.getlist('image_subactivity')

            report_for_id = Report.objects.last()

            id_rep = report_for_id.id

            repor = Report.objects.get(pk = id_rep)
            obser = 0
            
            for i in image:

                subactividad_img = SubActivity.objects.get(id=int(subactivity_image[obser]))     
                
                imagen = ImgReport(
                    image = i,
                    description = observation[obser]
                    )

                imagen.save()

                imagereport = ReportImage(
                    report = repor,
                    image = imagen,
                    subactivity = subactividad_img
                    )

                imagereport.save()

                obser = obser + 1

            if id_rep > 0:
                    
                data = {
                    'submitted': 1,
                    'id_report': id_rep
                    }

                create_pdf(id_rep)

            else:
                data = {
                    'submitted': 0
                    }

            return JsonResponse(data)

    else:
        return HttpResponse("<h2>No es posible agregar el reporte, verifique los datos correctamente</h2>")

### FORMULARIO DE BUSCAR REPORTES

@login_required(login_url="login")
def searchactivities(request):

    #Objetos a mostrar
    apis = API.objects.all()
    contracts = Contract.objects.all()
    specialties = Specialty.objects.all()
    references = Reference.objects.all()
    activities = Activity.objects.all()
    subactivities = SubActivity.objects.all()
    equipments = Equipment.objects.all()
    nonconformities = NonConformity.objects.all()
    followings = Following.objects.all()
    preconditions = Precondition.objects.all()
    measures = Measure.objects.all()
    activitytypes = ActivityType.objects.all()

    assert isinstance(request, HttpRequest)
    return render(
        request,
        'form-report/searchactivities.html',
        {
            'title':'Búsqueda de Actividades y Subactividades',
            'year':datetime.now().year,
            'apis': apis,
            'contracts': contracts,
            'specialties': specialties,
            'references': references,
            'activities': activities,
            'subactivities': subactivities,
            'equipments': equipments,
            'nonconformities': nonconformities,
            'followings': followings,
            'preconditions': preconditions,
            'measures': measures,
            'activitytypes': activitytypes,
        }
    )

search = []
@login_required(login_url="login")
def busquedaactividades(request, *args, **kwargs):


    try:
       if request.method == 'GET':

            search.clear()

            listar = request.GET.getlist('listar[]')
            busqueda = json.loads(listar[0])

            contrato = Contract.objects.get(pk=int(busqueda[0]['id_contrato']))
            api = API.objects.get(pk=int(busqueda[0]['id_api']))

            for i in busqueda:

                historicos = Historical.objects.filter(activity=int(i['id_actividad'])).values('subactivity',
                                                                                                'activity', 
                                                                                                'measure', 
                                                                                                'activitytype',
                                                                                                'subactivity_no_program',
                                                                                                ).order_by('subactivity').annotate(totalacumulado=Sum('real_amount'))
            #search = []

            for i in historicos:

                #historic = Historical.objects.filter(subactivity = int(i['subactivity'])).latest('inspection_date')
                historic = Historical.objects.filter(subactivity = int(i['subactivity']))

                act = Activity.objects.get(id = int(i['activity']))
                sub = SubActivity.objects.get(id = int(i['subactivity']))
                acttype = ActivityType.objects.get(id = int(i['activitytype']))
                measure = Measure.objects.get(id = int(i['measure']))
                
                for historico in historic:
                        noconfor = NonConformity.objects.get(id = int(historico.nonconformity_id))
                        report = Report.objects.get(id = int(historico.report_id))
                        segui = Following.objects.get(id = int(report.following_id))

                        dias = sub.finish_date - sub.start_date 
                        diaspro = dias.days
                        if diaspro == 0:
                            diaspro = 1

                        #dias_a = (str(report.inspection_date.day) + "/" + str(report.inspection_date.month) + "/" + str(report.inspection_date.year)) - (str(sub.start_date ) + "/" + str(sub.start_date ) + "/" + str(sub.start_date ))
                        dias_a = report.inspection_date.date() - sub.start_date
                        dias_acum = dias_a.days

                        if historico.activitytype.id == 2:
                            nombre_subactivity = historico.subactivity_no_program + " (No Programada)"
                            dias_acum = 1

                        else:
                            nombre_subactivity = sub.subactivity_name

                        if dias_acum < 1:
                            dias_acum = 1

                        search.append({
                            'id_historico': historico.id,
                            'id_contrato': contrato.id,
                            'id_api': api.id,
                            'contrato': contrato.contract_name, 
                            'contrato_numero': contrato.contract_number, 
                            'api': api.api_number,
                            'id_actividad': i['activity'],
                            'reporte': report.id,
                            'fecha_reporte': str(report.inspection_date.day) + "/"
                                             + str(report.inspection_date.month) + "/"
                                             + str(report.inspection_date.year),
                            'avance_diario': str(historico.real_amount),
                            'ref_dia': str(round(sub.average_amount/diaspro, 2)),
                            'dias': diaspro,
                            'actividad': act.activity_name,
                            'id_subactividad': i['subactivity'],
                            'subactividad': nombre_subactivity,
                            'total_estimado': sub.average_amount,
                            'id_medida': i['measure'],
                            'medida': measure.measure_name,
                            'total_acumulado': historico.difference,
                            'dias_acumulado': dias_acum,
                            'fecha_inicio': str(sub.start_date.day) + "/"
                                     + str(sub.start_date.month) + "/"
                                     + str(sub.start_date.year),
                            'fecha_termino': str(sub.finish_date.day) + "/"
                                     + str(sub.finish_date.month) + "/"
                                     + str(sub.finish_date.year),
                            'id_seguimiento': segui.id,
                            'seguimiento': segui.following_name,
                            'id_tipo_sub': acttype.id,
                            'tipo_subactividad': historico.activitytype.activity_type_name,
                            'id_no_conformidad': noconfor.id,
                            'no_conformidad': noconfor.nonconformity_name
                            })
            
            if int(busqueda[0]['id_subactividad']) != 0:
                        
                large = len(search)  
                inc = 0
                while inc < large:
                
                    if (search[inc]['id_subactividad'] != int(busqueda[0]['id_subactividad'])):

                        search.pop(inc)

                        inc = inc
                        large -= 1
                    else:
                        inc += 1

            if int(busqueda[0]['id_seguimiento']) != 0:
                        
                large = len(search)  
                inc = 0
                while inc < large:
                
                    if (search[inc]['id_seguimiento'] != int((busqueda[0]['id_seguimiento']))):

                        search.pop(inc)

                        inc = inc
                        large -= 1
                    else:
                        inc += 1

            if int(busqueda[0]['id_tipoactividad']) != 0:
                        
                large = len(search)  
                inc = 0
                while inc < large:
                
                    if (search[inc]['id_tipo_sub'] != int((busqueda[0]['id_tipoactividad']))):

                        search.pop(inc)

                        inc = inc
                        large -= 1
                    else:
                        inc += 1

            if int(busqueda[0]['id_conformidad']) != 0:
                        
                large = len(search)  
                inc = 0
                while inc < large:
                
                    if (search[inc]['id_no_conformidad'] != int((busqueda[0]['id_conformidad']))):

                        search.pop(inc)

                        inc = inc
                        large -= 1
                    else:
                        inc += 1

            if busqueda[0]['fecha_inicio'] != "" and busqueda[0]['fecha_termino'] != "":
                        
                large = len(search)  
                inc = 0

                startdate = busqueda[inc]['fecha_inicio']
                fecha_st = datetime.strptime(startdate, '%d/%m/%Y')

                finishdate = busqueda[inc]['fecha_termino']
                fecha_ed = datetime.strptime(finishdate, '%d/%m/%Y')

                while inc < large:
                    
                    report_date = search[inc]['fecha_reporte']
                    fecha_rp = datetime.strptime(report_date, '%d/%m/%Y')

                    if fecha_rp < fecha_st or fecha_rp > fecha_ed:

                        search.pop(inc)

                        inc = inc
                        large -= 1
                    else:
                        inc += 1

    except Exception as e:
        respuesta = e

    #paginator = Paginator(search, 10)

    #page = request.GET.get('page')
    #page_search = paginator.get_page(page)

    return JsonResponse(search, safe=False)

#FUNCION PARA AJUSTAR EL ANCHO DE LAS COLUMNAS DE LOS EXCEL
def exceladjust(excel):
    for col in excel.columns:
        max_lenght = 0
        col_name = re.findall('\w\d', str(col[0]))
        col_name = col_name[0]
        col_name = re.findall('\w', str(col_name))[0]
        for cell in col:
            try:
                if len(str(cell.value)) > max_lenght:
                    max_lenght = len(cell.value)
            except:
                pass
        adjusted_width = (max_lenght+2)
        excel.column_dimensions[col_name].width = adjusted_width

def dicttolist(lista, excels):

    for l in lista:

        listkey = l.values()
        listkey = list(listkey)

        excels.append(listkey)    

    return excels

def transformtotable(headers, infoinserted, excel, name):

    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
                'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT']

    ancho = len(headers) - 1
    largo = len(infoinserted) + 1

    tab = TableExcel(displayName="Table"+ name, ref="A1:"+ alphabet[ancho] + str(largo))

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style

    excel.add_table(tab)

    return excel

@csrf_exempt
@login_required(login_url="login")
def deletehistoric(request, *args, **kwargs):

    try:
        if request.method == 'POST':

            id_hist = request.POST.get('id')

            Historical.objects.filter(pk=id_hist).delete()
            
    except Exception as e:
        response = e
            
#FUNCION PARA DESCARGAR EL EXCEL AL REALIZAR UNA BÚSQUEDA DE ACTIVIDAD Y SUBACTIVIDAD
@login_required(login_url="login")
def downloadexcelsearch(request, id):

    wb = Workbook()

    headers = ['ID HISTORICO','ID CONTRATO', 'ID API', 'CONTRATO', 'CONTRATO NUMERO',
                'API', 'ID ACTIVIDAD', 'REPORTE', 'FECHA REPORTE', 'AVANCE DIARIO','REFERENCIA DÍA','DÍAS PROGRAMADO','ACTIVIDAD', 'ID SUBACTIVIDAD', 
                'SUBACTIVIDAD', 'TOTAL PROGRAMADO',
                'ID MEDIDA', 'MEDIDA', 'AVANCE DIARIO ACUMULADO','AVANCE DÍAS ACUMULADO', 'FECHA INICIO', 'FECHA TERMINO',
                'ID SEGUIMIENTO', 'SEGUIMIENTO', 'ID TIPO SUBACTIVIDAD',
                'TIPO SUBACTIVIDAD', 'ID NO CONFORMIDAD', 'NO CONFORMIDAD']

    headersequipos = ['NOMBRE DEL EQUIPO', 'CANTIDAD', 'DOTACION DIRECTA', 'DOTACION INDIRECTA', 'DOTACION REFERENCIAL', 'NOMBRE ACTIVIDAD']

    headersreference = ['NOMBRE REFERENCIA', 'DESCRIPCION', 'NUMERO REPORTE', 'FECHA DE REPORTE']

    headersforall = ['ID CONTRATO', 'CONTRATO', 'NUMERO CONTRATO', 'EMPRESA CONTRATISTA', 'RUT EMPRESA', 'JEFE DE PROYECTO', 'CORREO EMPRESA', 'TELEFONO EMPRESA',
                     'FECHA INICIO CONTRATO', 'FECHA TERMINO CONTRATO', 
                    'API','NUMERO DE API','NUMERO DE REPORTE', 'FECHA DE REPORTE','ACTIVIDAD', 'INICIO ACTIVIDAD',
                    'TERMINO ACTIVIDAD', 'SUBACTIVIDAD','CANTIDAD A LOGRAR', 'PROMEDIO HH',
                    'INICIO SUBACTIVIDAD', 'TERMINO SUBACTIVIDAD', 'UNIDAD', 'CANTIDAD REAL', 'TIPO DE ACTIVIDAD',
                    'PRECONDICION','¿CONFORMIDAD?',
                    'USUARIO', 'ESPECIALIDAD', 'DESVIACION DETECTADA',
                    'PLAN DE ACCION', 'SEGUIMIENTO']    

    #EXTRACCIÓN DE INFORMACIÓN DE ACTIVIDADES Y SUBACTIVIDADES PROGRAMADAS DE UN CONTRATO

    contrato_id = search[0]['id_contrato']
    api_id = search[0]['id_api']
    print(id)

    contrato = Contract.objects.filter(id=int(contrato_id))
    api = API.objects.filter(id=int(api_id))
    actividades = Activity.objects.filter(contract=int(contrato_id))
    subs = []
    for a in actividades:
         
        subactividades = SubActivity.objects.filter(activity=int(a.id))

        for s in subactividades:

            subs.append({
                    'id_subactividad': s.id
                })

    histprogram = []
    historicalref = []

    for su in subs:

        historicos = Historical.objects.filter(subactivity=int(su['id_subactividad']))

        for h in historicos:

            follow = Following.objects.filter(id=int(h.report.following_id))
            histprogram.append({
                    'id_contrato': contrato[0].id,
                    'contrato': contrato[0].contract_name,
                    'numero_contrato': contrato[0].contract_number,
                    'empresa': contrato[0].enterprise,
                    'rut': contrato[0].rut,
                    'jefe': contrato[0].project_boss,
                    'correo': contrato[0].email,
                    'telefono': contrato[0].cellphone,
                    'inicio_contrato': contrato[0].start_date,
                    'termino_contrato': contrato[0].finish_date,
                    'api': api[0].project_name,
                    'numero_api': api[0].api_number,
                    'numero_reporte': h.report_id,
                    'fecha_reporte': str(h.report.inspection_date.day) + "/"
                                     + str(h.report.inspection_date.month) + "/"
                                     + str(h.report.inspection_date.year),
                    #'id_actividad': h.activity_id,
                    'actividad': h.activity.activity_name,
                    'inicio_actividad': h.activity.start_date,
                    'termino_actividad': h.activity.finish_date,
                    #'id_subactividad': h.subactivity_id,
                    'nombre_subactividad': h.subactivity.subactivity_name,
                    'cantidad_a_realizar': h.subactivity.average_amount,
                    'promedio_hh': h.subactivity.average_hh,
                    'inicio_subactivity': str(h.subactivity.start_date.day) + "/"
                                     + str(h.subactivity.start_date.month) + "/"
                                     + str(h.subactivity.start_date.year),
                    'termino_subactivity': str(h.subactivity.finish_date.day) + "/"
                                     + str(h.subactivity.finish_date.month) + "/"
                                     + str(h.subactivity.finish_date.year),
                    'unidad': h.measure.measure_name,
                    #'id_historico': h.id,
                    'cantidad_real': h.real_amount,
                    'tipo_actividad': h.activitytype.activity_type_name,
                    'precondicion': h.precondition.precondition_name,
                    'nonconformity': h.nonconformity.nonconformity_name,
                    'usuario': h.user.first_name + " " + h.user.last_name,
                    'especialidad': h.specialty.specialty_name,
                    'desviacion': h.report.deviation_detected,
                    'plan_de_accion': h.report.action_plan,
                    'seguimiento': follow[0].following_name

                })
            histrefe = HistoricalReference.objects.filter(report=int(h.report_id))

            for hf in histrefe:

                historicalref.append({
                    'nombre_referencia': hf.reference.reference_name,
                    'descripcion': hf.description,
                    'numero_reporte': hf.report_id,
                    'fecha_reporte': str(hf.report.inspection_date.day) + "/"
                                     + str(hf.report.inspection_date.month) + "/"
                                     + str(hf.report.inspection_date.year)
                    })
    
    equiposcantidad = []

    for act in actividades:
        equiamo = EquipmentAmount.objects.filter(activity=int(act.id))
        for equi in equiamo:
            equiposcantidad.append({
                'nombre_equipo':equi.equipment.equipment_name,
                'cantidad': equi.equipment_amount,
                'dotacion_directa': equi.direct_endowment,
                'dotacion_indirecta':equi.indirect_endowment,
                'dotacion_referencial':equi.direct_reference,
                'actividad':equi.activity.activity_name
                })



    wb.remove(wb.active)

    page = wb.create_sheet('BÚSQUEDA ESPECIFICA')
    page2 = wb.create_sheet('HISTORICO DE REPORTES')
    page3 = wb.create_sheet('REFERENCIAS')
    page4 = wb.create_sheet('EQUIPOS')

    page.append(headers)
    page2.append(headersforall)
    page3.append(headersreference)
    page4.append(headersequipos)

    page = dicttolist(search, page)
    page2 = dicttolist(histprogram, page2)
    page3 = dicttolist(historicalref, page3)
    page4 = dicttolist(equiposcantidad, page4)

    page = transformtotable(headers, search, page, "Búsqueda")
    page2 = transformtotable(headersforall, histprogram, page2,"Historico")
    page3 = transformtotable(headersreference, historicalref, page3, "Referencias")
    page4 = transformtotable(headersequipos, equiposcantidad, page4, "Equipos")

    page = exceladjust(page)
    page2 = exceladjust(page2)
    page3 = exceladjust(page3)
    page4 = exceladjust(page4)

    now = datetime.now()
    format = now.strftime('%d/%m/%Y')
    nombre_archivo = "Reporte-de-"+ request.user.first_name + "_" + request.user.last_name + "("+str(format)+").xlsx"
    response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="'+ nombre_archivo +'"'
    return response

    wb.save(response)

    return response

# FORMULARIO PARA LA CARGA DE DATOS
@login_required(login_url="login")
@user_passes_test(lambda u: u.groups.filter(name='P&C').exists(), login_url=home)
def loadfiles(request):
    assert isinstance(request, HttpRequest)
    return render(
        request,
        'form-report/loadfiles.html',
        {
            'title':'Cargar Datos',
            'year':datetime.now().year,
        }
    )

def logopdf(c, h):
    url_img = settings.STATIC_ROOT + "/app/img/logo.png" 
    im_logo = Image(url_img, width=91, height=60)
    im_logo.drawOn(c, 20, h-68)


#Validación de datos correctos en la carga del excel.
def verifydata(data, name, val, valnum):
    
    val = val
    valnum = valnum
    real_val = 0

    if valnum == 2:

        real_val = 2

    if data == None or data == "":

        val = val + "El campo " + name + " se encuentra vacío. "
        valnum = 2

    else: 

        if name == "Número contrato" or name == "Total programado" or name == "Numero de API":

            if str(data).isdigit():
                valnum = 1

            else:
                val = val + "El campo " + name + " contiene letras o caracteres que no son númericos. "
                valnum = 2

        if name == "Referencia Diaria":

            if (type(data) is float) or (type(data) is int):

                valnum = 0

            else:

                val = val + "El campo " + name + " no es númerico o decimal. "
                valnum = 2

        if name == "Inicio de API" or name == "Termino de API" or name == "Inicio de Contrato" or name == "Termino de Contrato" or name == "Inicio de Actividad" or name == "Termino de Actividad" or name == "Fecha inicio programado" or name == "Fecha termino programado" or name == "Fecha Inicio proyectado" or name == "Fecha termino proyectado":

            if type(data) is datetime:

                valnum = 0

            else:

                val = val + "El campo " + name + " no es un valor de tipo fecha. "
                valnum = 2



    if real_val == 2:

        return val, real_val

    else:

        return val, valnum



#CARGA DE DATOS A TRAVES DE UN EXCEL
@csrf_exempt
@login_required(login_url="login")
def readexcel(request):
    try:
        if request.method == 'POST':
            form = UploadFileForm(request.POST, request.FILES)
            rd = form.is_valid() 
            if form.is_valid():

                wb = load_workbook(filename=request.FILES['file'].file, data_only=True)

                sheet_obj = wb.active
                m_row = sheet_obj.max_row
            
                subactividades = []

                act_index = 0
                isub = 0

                for i in range(3, m_row + 1):

                    if (sheet_obj.cell(row = i, column = 1).value != None) or (sheet_obj.cell(row = i, column = 1).value != "NUMERO API"):

                        val_num = 0
                        validacion = ""
                        api_val = ""
                        con_val = ""
                        act_val = ""
                        sub_val = ""

                        numero_api = sheet_obj.cell(row = i, column = 1)
                        #nombre_proyecto = sheet_obj.cell(row = i, column = 2)
                        inicio_api = sheet_obj.cell(row = i, column = 2)
                        termino_api = sheet_obj.cell(row = i, column = 3)
                        #contrato = sheet_obj.cell(row = i, column = 5)
                        numero_contrato = sheet_obj.cell(row = i, column = 4)
                        #empresa = sheet_obj.cell(row = i, column = 6)
                        #jefe = sheet_obj.cell(row = i, column = 7)
                        #correo = sheet_obj.cell(row = i, column = 8)
                        #tel_contacto = sheet_obj.cell(row = i, column = 9)
                        #rut_empresa = sheet_obj.cell(row = i, column = 10)
                        inicio_contrato = sheet_obj.cell(row = i, column = 5)
                        termino_contrato = sheet_obj.cell(row = i, column = 6)
                        actividad = sheet_obj.cell(row = i, column = 7)
                        inicio_actividad = sheet_obj.cell(row = i, column = 8)
                        termino_actividad = sheet_obj.cell(row = i, column = 9)
                        subactividad = sheet_obj.cell(row = i, column = 10)

                        dias_programado = sheet_obj.cell(row = i, column = 11)
                        fecha_inicio_promagado = sheet_obj.cell(row = i, column = 12)
                        fecha_termino_promagado = sheet_obj.cell(row = i, column = 13)

                        dias_proyectado = sheet_obj.cell(row = i, column = 14)
                        fecha_inicio_proyectado = sheet_obj.cell(row = i, column = 15)
                        fecha_termino_proyectado = sheet_obj.cell(row = i, column = 16)

                        total = sheet_obj.cell(row = i, column = 17)
                        diaria = sheet_obj.cell(row = i, column = 18)
                        unidad = sheet_obj.cell(row = i, column = 19)


                        #Validaciones primarias (Verificación si existen aquellos datos)

                        #Uso de la variable num_save:
                        #- num_save = 1 : se agregara desde la api en adelante
                        #- num_save = 2 : se agregara desde el contrato en adelante
                        #- num_save = 3 : se agregara desde la actividad en adelante
                        #- num_save = 4 : se agregara desde la subactividad en adelante

                        if (numero_api.value is not None) and str(numero_api.value).isdigit():

                            api_val = API.objects.filter(api_number=int(numero_api.value))

                            if len(api_val) != 0:

                                num_save = 1

                                if (numero_contrato.value is not None) and str(numero_contrato.value).isdigit():
                                
                                    con_val = Contract.objects.filter(contract_number=int(numero_contrato.value))

                                    if len(con_val) != 0:

                                        num_save = 2

                                        if actividad.value is not None:

                                            act_val = Activity.objects.filter(activity_name=actividad.value.strip())
                                        
                                            if len(act_val) != 0:

                                                num_save = 3

                                                if subactividad is not None:

                                                    sub_val = SubActivity.objects.filter(subactivity_name=subactividad.value.strip())

                                                    if len(sub_val) != 0:

                                                        num_save = 4

                                                        val_num = 1
                                                        validacion = "SubActividad encontrada, se actualizara su total programado."
     
                                                    else:

                                                        validacion = "SubActividad no existe, por lo tanto se agregara la subactividad."
                                                        val_num = 1

                                                else:
                                                    
                                                    validacion = "SubActividad se encuentra vacía."
                                                    val_num = 2

                                            else:

                                                validacion = "Actividad no existe, por lo tanto se agregara con todas las subactividades."
                                                val_num = 1

                                        else:

                                            validacion = "Actividad se encuentra vacía."
                                            val_num = 2

                                    else:

                                        validacion = "Número de Contrato no existe, por lo tanto se agregara todo su contenido."
                                        val_num = 1

                                else:

                                    validacion = "Número de Contrato se encuentra vacío o no es numérico."
                                    val_num = 2
                            
                            else:
                            
                                validacion = "Número de API no existe en los registros, por lo tanto se guardara todos los datos consiguientes."
                                val_num = 1

                        else: 

                            validacion = "Número de API se encuentra vacío o no es numérico."
                            val_num = 2

                        #Validaciones secundarias (Verificación de datos a ingresar sean correctos)

                        if len(api_val) == 0:

                            validacion, val_num = verifydata(numero_api.value, "Numero de API", validacion, val_num)
                            validacion, val_num = verifydata(inicio_api.value, "Inicio de API", validacion, val_num)
                            validacion, val_num = verifydata(termino_api.value, "Termino de API", validacion, val_num)
                            #validacion, val_num = verifydata(nombre_proyecto.value, "Nombre Proyecto", validacion, val_num)

                        if len(con_val) == 0: 

                            validacion, val_num = verifydata(numero_contrato.value, "Número contrato", validacion, val_num)
                            #validacion, val_num = verifydata(contrato.value, "Contrato", validacion, val_num)
                            validacion, val_num = verifydata(termino_contrato.value, "Termino de Contrato", validacion, val_num)
                            validacion, val_num = verifydata(inicio_contrato.value, "Inicio de Contrato", validacion, val_num)
                            #validacion, val_num = verifydata(empresa.value, "Empresa", validacion, val_num)
                            #validacion, val_num = verifydata(tel_contacto.value, "Telefono de Contacto", validacion, val_num)
                            #validacion, val_num = verifydata(rut_empresa.value, "RUT Empresa", validacion, val_num)

                        if len(act_val) == 0:

                            validacion, val_num = verifydata(inicio_actividad.value, "Inicio de Actividad", validacion, val_num)
                            validacion, val_num = verifydata(actividad.value, "Actividad", validacion, val_num)  
                            validacion, val_num = verifydata(termino_actividad.value, "Termino de Actividad", validacion, val_num)

                        if len(sub_val) == 0:

                            validacion, val_num = verifydata(subactividad.value, "Sub Actividad", validacion, val_num)
                            validacion, val_num = verifydata(dias_programado.value, "Días programado", validacion, val_num)
                            validacion, val_num = verifydata(fecha_inicio_promagado.value, "Fecha inicio programado", validacion, val_num)
                            validacion, val_num = verifydata(fecha_termino_promagado.value, "Fecha termino programado", validacion, val_num)
                            validacion, val_num = verifydata(dias_proyectado.value, "Dias proyectado", validacion, val_num)
                            validacion, val_num = verifydata(fecha_inicio_proyectado.value, "Fecha Inicio proyectado", validacion, val_num)
                            validacion, val_num = verifydata(fecha_termino_proyectado.value, "Fecha termino proyectado", validacion, val_num)
                            validacion, val_num = verifydata(diaria.value, "Referencia Diaria", validacion, val_num)
                            validacion, val_num = verifydata(unidad.value, "Unidad", validacion, val_num)
                            validacion, val_num = verifydata(total.value, "Total programado", validacion, val_num)

                        if val_num == 0:

                            val_name = "good"

                        if val_num == 1:

                            val_name = "careful"

                        if val_num == 2:

                            val_name = "danger"

                        #or (type(diaria.value) is not float)
                        if (diaria.value is not None):
                            ref_dia = round(diaria.value, 3)
                        else:
                            ref_dia = diaria.value

                        subactividades.append({
                            'numero_api': numero_api.value,
                            #'nombre_proyecto': nombre_proyecto.value,
                            'inicio_api': inicio_api.value,
                            'termino_api': termino_api.value,
                            #'contrato': numero_contrato.value,
                            'numero_contrato': numero_contrato.value,
                            #'empresa': empresa.value,
                            #'tel_contacto': tel_contacto.value,
                            #'rut_empresa': rut_empresa.value,
                            #'jefe_proyecto': jefe.value,
                            #'correo': correo.value,
                            'inicio_contrato': inicio_contrato.value,
                            'termino_contrato': termino_contrato.value,
                            'actividad': actividad.value,
                            'inicio_actividad': inicio_actividad.value,
                            'termino_actividad': termino_actividad.value,
                            'subactividad': subactividad.value,
                            'dias_programado': dias_programado.value,
                            'fecha_inicio_programada': fecha_inicio_promagado.value.date(),
                            'fecha_termino_programada': fecha_termino_promagado.value.date(),
                            'dias_proyectado': dias_proyectado.value,
                            'fecha_inicio_proyectado': fecha_inicio_proyectado.value.date(),
                            'fecha_termino_proyectado': fecha_termino_proyectado.value.date(),
                            'total_estimado': total.value,
                            'referencia_diaria': ref_dia,
                            'medida': unidad.value,
                            'obs_validacion': validacion,
                            'clase_validacion': val_name,
                            'numero_validacion': val_num,
                            #'validacion_guardar': num_save
                            })
                        print(subactividades[isub])

                        isub += 1

            return JsonResponse(subactividades, safe=False)
    except Exception as e:
        response = e

        return response

def savecontract(data, a):

    con = Contract.objects.filter(contract_number = int(data['numero_contrato'])).filter(api = a.id)

    if len(con) != 0:
        conies = con[0]
    else:
        conies = None

    if conies is None:
      
        contr = Contract(
                #contract_name = data['contrato'],
                contract_number = data['numero_contrato'],
                #enterprise = data['empresa'],
                #rut = data['rut_empresa'],
                #project_boss = data['jefe_proyecto'],
                #email = data['correo'],
                #cellphone = data['tel_contacto'],
                start_date = datetime.strptime(data['inicio_contrato'], "%Y-%m-%dT%H:%M:%S"),
                finish_date = datetime.strptime(data['termino_contrato'], "%Y-%m-%dT%H:%M:%S"),
                state = 1,
                api = a
            )

        contr.save()

        return contr

    else:
        contrato = con[0]
        return contrato

def saveapi(data):

    api = API.objects.filter(api_number = data['numero_api'])

    if len(api) != 0:
        apies = api[0]
    else:
        apies = None

    if apies is None:

        apis = API(
                api_number = data['numero_api'],
                #project_name = data['nombre_proyecto'].strip(),
                start_date = datetime.strptime(data['inicio_api'], "%Y-%m-%dT%H:%M:%S"),
                finish_date = datetime.strptime(data['termino_api'], "%Y-%m-%dT%H:%M:%S"),
                state = 1
            )

        apis.save()

        return apis

    else:
        apireal = api[0]
        return apireal

def saveactivity(data, c, a):

    act = Activity.objects.filter(activity_name = data['actividad'].strip()).filter(contract=c.id).filter(api=a.id)

    if len(act) != 0:
        acties = act[0]
    else:
        acties = None

    if acties is None:

        acts = Activity(
                activity_name = data['actividad'].strip(),
                start_date = datetime.strptime(data['inicio_actividad'], "%Y-%m-%dT%H:%M:%S"),
                finish_date = datetime.strptime(data['termino_actividad'], "%Y-%m-%dT%H:%M:%S"),
                state = 1,
                api = a,
                contract = c
            )

        acts.save()

        return acts

    else:
        actividad = act[0]
        return actividad

def savesubactivity(data, c, a, ac, m):

    subs = SubActivity.objects.filter(subactivity_name = data['subactividad'].strip()).filter(api=a.id).filter(contract=c.id).filter(activity=ac.id)
    if len(subs) != 0:
        subsies = subs[0]
    else:
        subsies = None

    if subsies is None:

        subsa = SubActivity(
                subactivity_name = data['subactividad'].strip(),
                average_amount = data['total_estimado'],
                average_hh = data['total_estimado'],
                start_date = datetime.strptime(data['fecha_inicio_programada'], "%Y-%m-%d"),
                finish_date = datetime.strptime(data['fecha_termino_programada'], "%Y-%m-%d"),
                days = data['dias_programado'],
                api = a,
                contract = c,
                activity = ac,
                measureunit = m,
                ref_day = data['referencia_diaria']
            )

        subsa.save()

        return subsa

    else:
        subs.update(average_amount=data['total_estimado'])
        subactividad = subs[0]
        return subactividad

def savemeasure(data):

    mea = Measure.objects.all()
    founded = 0

    for unit in mea:

        if unit.measure_name.lower() == data['medida'].lower():

            founded = 1
            un = unit

    if founded == 0:

        un = Measure(
                measure_name = data['medida']
            )

        un.save()

    return un

#SUBIDA DE DATOS DESDE UN EXCEL
@csrf_exempt
@login_required(login_url="login")
def submitdata(request, *args, **kwargs):

    if request.method == 'POST':

        subm = request.POST.getlist('data[]')
        submit = json.loads(subm[0])

        i = 1
        s = 0
        large = len(submit)
        #Verificación a través del validador incluido para averiguar si los datos en general estan correctos para agregar
        while i == 1:
            print (i)

            if int(submit[s]['numero_validacion']) ==  2:

                i = 0

            else:

                if (s + 1) == large:

                    i = 2

                else:

                    s+=1

        if i == 1 or i == 2:

            for sub in submit:

                proyecto = saveapi(sub)
                contrato = savecontract(sub, proyecto)
                actividad = saveactivity(sub, contrato, proyecto)
                medida = savemeasure(sub)
                subactividad = savesubactivity(sub, contrato, proyecto, actividad, medida)

                data = 1

            return JsonResponse(data, safe=False)

        else:

            data = 0
            return JsonResponse(data, safe=False)



def downloadstructure(request, id):
    
    if id == 'structure':
        file_path = settings.STATIC_ROOT + "/app/downloadexcel/EstructuraCargadeDatos.xlsx" 

    else: 
        file_path = settings.STATIC_ROOT + "/app/downloadexcel/EstructuradeEjemplo.xlsx" 

    if os.path.exists(file_path):
        with open(file_path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="application/vnd.ms-Excel")
            response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
            return response
    raise Http404

# DESCARGA DE PDF AL REGISTRAR UN INFORME DIARIO
@csrf_exempt
@login_required(login_url="login")
def downloadpdf(request):

    if request.method == 'POST':
        rep_id = request.POST['id']
        file_path = settings.MEDIA_ROOT + "/pdf_reports/REPORTE_N" +  rep_id + ".pdf"

        if os.path.exists(file_path):
            with open(file_path, 'rb') as fh:
                response = HttpResponse(fh.read(), content_type="application/pdf")
                response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                buffer = io.BytesIO()

                buffer.seek(0)
                pdf: bytes = buffer.getvalue()

                response.write(pdf)
                file_data = ContentFile(pdf)

                return response

        raise Http404


def create_pdf(id):

    id_rep = id

    reportehecho = Report.objects.filter(id=int(id_rep))
    historico = Historical.objects.filter(report=int(id_rep))
    subactividad = SubActivity.objects.filter(id=int(historico[0].subactivity_id))
    actividad = Activity.objects.filter(id=int(subactividad[0].activity_id))
    contrato = Contract.objects.filter(id=int(actividad[0].contract_id))
    api = API.objects.filter(id=int(contrato[0].api_id))
    hist_ref = HistoricalReference.objects.filter(report=id_rep)
    hist_equi = EquipmentAmount.objects.filter(report=id_rep)
    hist_img = ReportImage.objects.filter(report=id_rep)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment;filename=REPORTE N°'+ str(id_rep)+'.pdf'

    buffer = io.BytesIO()

    c = canvas.Canvas(
                      buffer,
                      pagesize=portrait(letter)
                      )
    width, height = portrait(letter) 
    stream = io.BytesIO()
    
    styles = getSampleStyleSheet()
    styleN = styles["BodyText"]
    styleN.alignment = TA_LEFT
    styleBH = styles["Normal"]
    styleBH.alignment = TA_CENTER

    #Cabezera PDF
    archivo_imagen = settings.STATIC_ROOT +'/app/img/logo.jpg'

    height_pdf = height
    url_img = settings.STATIC_ROOT + "/app/img/report_banners/InformeBanner.png" 
    im_banner = Image(url_img, width=600, height=115)
    im_banner.drawOn(c, 5, 680)

    height_pdf -= 115

    c.setFont('Helvetica', 18)
    title_report = stringWidth("REPORTE N°"+str(id_rep), 'Helvetica', 18)

    c.drawString((width/2)-(title_report/2), 670,"REPORTE N°"+str(id_rep))

    c.setFont('Helvetica', 10)
    c.drawString(25, 635,"Autor:")
    c.drawString(130, 635, historico[0].user.first_name + " " + 
                                     historico[0].user.last_name)

    c.drawString(25, 655,"Fecha de Inspección: ")
    c.drawString(130, 655, str(historico[0].report.inspection_date.day) + "/"
                                     + str(historico[0].report.inspection_date.month) + "/"
                                     + str(historico[0].report.inspection_date.year))

    height_pdf -= 40
    #Primera Parte PDF: Antecedentes Generales.
    url_img = settings.STATIC_ROOT + "/app/img/report_banners/AntecedentesBanner.png" 
    im_banner = Image(url_img, width=600, height=47)
    im_banner.drawOn(c, 5, 590)

    height_pdf -= 47

    c.setFont('Helvetica', 10)
    c.drawString(25, 585,"N° de API: " + api[0].api_number)
    c.drawString(160, 585,"N° de Contrato: " + str(contrato[0].contract_number))
    c.drawString(320, 585,"Nombre de Proyecto: " + api[0].project_name)
    c.drawString(25, 565,"Empresa Contratista: " + contrato[0].enterprise)
    c.drawString(25, 545,"Fecha de Inicio: "+ str(contrato[0].start_date.day) + "/"
                                     + str(contrato[0].start_date.month) + "/"
                                     + str(contrato[0].start_date.year))
    c.drawString(160, 545,"Fecha de Termino: "+ str(contrato[0].finish_date.day) + "/"
                                     + str(contrato[0].finish_date.month) + "/"
                                     + str(contrato[0].finish_date.year))
    c.drawString(320, 545,"Especialidad: " + historico[0].specialty.specialty_name)
    c.drawString(25, 525,"Referencias Utilizadas: ")

    height_pdf -= 100

    cabezera_referencia = ['Nombre Referencia', 'Descripción']
    data = [cabezera_referencia] 
    for hr in hist_ref:
        data += [[hr.reference.reference_name, hr.description]]

    t=Table(data)
    t.setStyle(TableStyle([ ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                            ('FONTSIZE', (0,0), (-1, -1), 10),
                            ('LINEABOVE', (0,1), (-1,1), 0.5, colors.black),
                            ('LINEBEFORE', (2,0), (2,-1), 0.5, colors.black),
                            ('ALIGN', (0,0), (1,-1), 'LEFT'),
                            ('ALIGN', (2,0), (-1,-1), 'RIGHT')]))

    t.wrapOn(c, width - 400, height)
    w, h = t.wrap(100, 100)
    t.drawOn(c, 190, height - (h + 265), 0)

    pagina = 1
    c.drawString(250, 20,"Página " + str(pagina))

    #Segunda Parte PDF: Resultados
    height_pdf -= (h + 20)
    url_img = settings.STATIC_ROOT + "/app/img/report_banners/ResultadoBanner.png" 
    im_banner = Image(url_img, width=600, height=47)
    im_banner.drawOn(c, 5, height_pdf)

    height_pdf -= 20 

    title_report = stringWidth(reportehecho[0].following.following_name, 'Helvetica', 10)
    w_draw = (width/2)-(title_report/2)
    c.drawString(w_draw, height_pdf + 2, reportehecho[0].following.following_name)

    url_img = settings.STATIC_ROOT + "/app/img/report_banners/finish.png" 
    im_banner = Image(url_img, width=15, height=15)
    im_banner.drawOn(c, w_draw - 20, height_pdf)

    #Cuarta Parte: Recursos EECC En Terreno

    height_pdf -= 47

    url_img = settings.STATIC_ROOT + "/app/img/report_banners/RecursosBanner.png" 
    im_banner = Image(url_img, width=600, height=47)
    im_banner.drawOn(c, 5, height_pdf)
    
    n_actividad = Paragraph('''<b>ACTIVIDAD</b>''', styleBH)
    n_equipo = Paragraph('''<b>EQUIPO</b>''', styleBH)
    n_cantidad = Paragraph('''<b>CANTIDAD</b>''', styleBH)
    n_d_directa = Paragraph('''<b>DOT. DIRECTA</b>''', styleBH)
    n_d_referencial = Paragraph('''<b>DOT. REFERENCIAL</b>''', styleBH)
    n_d_indirecta = Paragraph('''<b>DOT. INDIRECTA</b>''', styleBH)

    equipos = []

    equipos = [[n_actividad, n_equipo, n_cantidad, n_d_directa, n_d_referencial, n_d_indirecta]]

    for equi in hist_equi:
        equipos += [[Paragraph(equi.activity.activity_name , styleN),
                     Paragraph(equi.equipment.equipment_name, styleN),
                     Paragraph(str(equi.equipment_amount), styleN),
                     Paragraph(str(equi.direct_endowment), styleN),
                     Paragraph(str(equi.direct_reference), styleN),
                     Paragraph(str(equi.indirect_endowment), styleN)]]
    
    t=Table(equipos, colWidths=3.3 * cm)
    t.setStyle(TableStyle([ ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                            ('FONTSIZE', (0,0), (-1, -1), 10),
                            ('LINEABOVE', (0,1), (-1,1), 0.5, colors.black),
                            ('LINEBEFORE', (2,0), (2,-1), 0.5, colors.black),
                            ('ALIGN', (0,0), (1,-1), 'LEFT'),
                            ('ALIGN', (2,0), (-1,-1), 'RIGHT')]))

    t.wrapOn(c, width, height)
    w, h = t.wrap(100, 100)
    height_pdf -= h
    t.drawOn(c, 25, height_pdf, 0)

    #Tercera Parte PDF: Control de Avance por Actividad según cronograma.
    height_pdf = height
    height_pdf -= 115
    c.showPage()
    pagina += 1
    c.drawString(250, 20,"Página " + str(pagina))
    logopdf(c, height)
    url_img = settings.STATIC_ROOT + "/app/img/report_banners/AvanceBanner.png" 
    im_banner = Image(url_img, width=600, height=47)
    im_banner.drawOn(c, 5, height_pdf)

    n_activity = Paragraph('''<b>ACTIVIDAD</b>''', styleBH)
    n_subactivity = Paragraph('''<b>SUB ACTIVIDAD</b>''', styleBH)
    n_cantidad = Paragraph('''<b>AVANCE DIARIO</b>''', styleBH)
    n_unidad = Paragraph('''<b>UNIDAD</b>''', styleBH)
    n_programado = Paragraph('''<b>PROGRAMADO</b>''', styleBH)
    n_refdia = Paragraph('''<b>REF. DIA</b>''', styleBH)
    n_acumulado = Paragraph('''<b>AVANCE REAL ACUMULADO</b>''', styleBH)
    n_cumpli = Paragraph('''<b>CUMPLIMIENTO</b>''', styleBH)
    n_causa = Paragraph('''<b>CAUSA NO CUMPLIMIENTO</b>''', styleBH)
    n_tipo_act = Paragraph('''<b>TIPO ACTIVIDAD</b>''', styleBH)

    
    actividades = [[n_activity, n_subactivity, n_cantidad, n_unidad, n_programado, n_refdia, n_acumulado, n_cumpli, n_causa, n_tipo_act]]
    c.setFont('Helvetica', 10)
    inc_hist = 0

    for acts in historico:

        dias = acts.subactivity.finish_date - acts.subactivity.start_date 
        diaspro = dias.days

        if diaspro == 0:
            diaspro = 1

        if acts.nonconformity_id == 5:
            cumplimiento = 'Si'
            causa = 'Cumplida'
        else:
            cumplimiento = 'No'
            causa = acts.nonconformity.nonconformity_name

        if acts.activitytype_id == 1:
            activ_total = Historical.objects.values('subactivity_id').filter(subactivity=int(acts.subactivity_id)).annotate(total=Sum('real_amount'))
            actividades += [[Paragraph(acts.activity.activity_name, styleN), 
                                Paragraph(acts.subactivity.subactivity_name, styleN),
                                Paragraph(str(acts.real_amount), styleN), 
                                Paragraph(acts.measure.measure_name, styleN) ,
                                Paragraph(str(acts.subactivity.average_amount), styleN), 
                                Paragraph(str(round(acts.subactivity.average_amount/diaspro, 2)), styleN),
                                Paragraph(str(activ_total[0]['total']), styleN),
                                Paragraph(cumplimiento, styleN),
                                Paragraph(causa, styleN),
                                Paragraph(acts.activitytype.activity_type_name, styleN)]]
            inc_hist += 1

        else:
            activ_total = 0
            actividades += [[Paragraph(acts.activity.activity_name, styleN), 
                                Paragraph(acts.subactivity.subactivity_name, styleN),
                                Paragraph(str(acts.real_amount)), 
                                Paragraph(acts.measure.measure_name), 
                                Paragraph(str(acts.no_program_total)), 
                                Paragraph(str(acts.no_program_refday)),
                                Paragraph(str(acts.no_program_total_acu)),
                                Paragraph(cumplimiento),
                                Paragraph(causa),
                                Paragraph(acts.activitytype.activity_type_name)]]

            inc_hist += 1

    t=Table(actividades, colWidths=2.05 * cm)
    t.setStyle(TableStyle([ ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
                            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
                            ('FONTSIZE', (0,0), (-1, -1), 10),
                            ('LINEABOVE', (0,1), (-1,1), 0.5, colors.black),
                            ('LINEBEFORE', (2,0), (2,-1), 0.5, colors.black),
                            ('ALIGN', (0,0), (1,-1), 'LEFT'),
                            ('ALIGN', (2,0), (-1,-1), 'RIGHT')]),
                            )


    t.wrapOn(c, width, height)
    w, h = t.wrap(100, 100)
    height_pdf -= h 
    t.drawOn(c, 15, height_pdf, 0)

    #Cuarta Parte: Imagenes
    c.showPage()

    height_pdf = height
    height_pdf -= 115
    logopdf(c, height)
    url_img = settings.STATIC_ROOT + "/app/img/report_banners/FotosBanner.png" 
    im_banner = Image(url_img, width=600, height=47)
    im_banner.drawOn(c, 5, height_pdf)

    c.setFont('Helvetica', 10)

    pagina += 1
    c.drawString(250, 20,"Página " + str(pagina))
    height_pdf -= 160

    inc_img = 0
    img_space = 40
    inc_page = 0

    img_len = len(hist_img)

    if img_len == 0:

        url_img = settings.STATIC_ROOT + "/app/img/report_banners/noimage.png" 
        im_banner = Image(url_img, width=100, height=100)
        im_banner.drawOn(c, (width/2)-50, height_pdf + 60)

        c.setFont('Helvetica', 15)
        width_text = stringWidth("No se registro imágenes", 'Helvetica', 15)
        c.drawString((width/2)-(width_text/2), height_pdf + 40, "No se registro imágenes")
    for img in hist_img:

        if inc_img > 2:
            inc_img = 0
            height_pdf -= 310
            img_space = 40 

            if inc_page > 5:
                height_pdf = height
                height_pdf -= 115

                c.showPage()
                logopdf(c, height)
                url_img = settings.STATIC_ROOT + "/app/img/report_banners/FotosBanner.png" 
                im_banner = Image(url_img, width=600, height=47)
                im_banner.drawOn(c, 5, height_pdf)
                height_pdf = height
                height_pdf -= 115

                
                c.setFont('Helvetica', 10)

                pagina += 1
                c.drawString(250, 20,"Página " + str(pagina))
                height_pdf -= 160

                inc_page = 0

        url_img = settings.MEDIA_ROOT + "/" + img.image.image.name
        #imgsize = utils.ImageReader(url_img) 
        #iw, ih = imgsize.getSize() 
        #aspect = iw/float(ih) 

        subact_name = "<b>SUB ACTIVIDAD: </b>" + img.subactivity.subactivity_name
        n_subactivity = Paragraph(subact_name, styleBH)

        obs_name = "<b>OBSERVACION: </b>" + img.image.description
        n_obs = Paragraph(obs_name, styleBH)

        w,h = n_subactivity.wrap(150,150)
        n_subactivity.drawOn(c, img_space, height_pdf - 60)

        w,h = n_obs.wrap(150,150)
        n_obs.drawOn(c, img_space, height_pdf - 150)

        im = Image(url_img, width=150, height=150, hAlign='CENTER')
        im.drawOn(c, img_space, height_pdf)

        inc_img += 1
        img_space += 170
        inc_page += 1

    #Quinta Parte: Observaciones generales
    c.setFont('Helvetica', 10)

    if inc_page > 2:
        c.showPage()
        logopdf(c, height)

        height_pdf = height
        height_pdf -= 115
        pagina += 1
        c.setFont('Helvetica', 10)
        c.drawString(250, 20,"Página " + str(pagina))
        c.drawString(150, 700,"OBSERVACIONES")
    else:
        height_pdf -= (47 + 150)

            
    url_img = settings.STATIC_ROOT + "/app/img/report_banners/ObservacionesBanner.png" 
    im_banner = Image(url_img, width=600, height=47)
    im_banner.drawOn(c, 5, height_pdf)

    if reportehecho[0].deviation_detected.isspace() or len(reportehecho[0].deviation_detected)==0:
        variable_report = "Ninguna"
    else: 
        variable_report = reportehecho[0].deviation_detected

    n_desviacion = Paragraph("<b>DESVIACIONES DETECTADAS</b> <br />" + variable_report, styleBH)
    w,h = n_desviacion.wrap(180,150)
    n_desviacion.drawOn(c, 40, height_pdf - 160)

    if reportehecho[0].action_plan.isspace() or len(reportehecho[0].action_plan)==0:
        variable_report = "Ninguna"
    else: 
        variable_report = reportehecho[0].action_plan

    n_plan = Paragraph("<b>PLAN DE ACCION</b> <br />" + variable_report, styleBH)
    w,h = n_plan.wrap(180,150)
    n_plan.drawOn(c, 230, height_pdf - 160)

    if reportehecho[0].evidence_obs.isspace() or len(reportehecho[0].evidence_obs)==0:
        variable_report = "Ninguna"
    else: 
        variable_report = reportehecho[0].evidence_obs

    n_obsimg = Paragraph("<b>OBSERVACIONES GENERALES DE FOTOGRAFIAS</b> <br />" + variable_report, styleBH)
    w,h = n_obsimg.wrap(180,150)
    n_obsimg.drawOn(c, 390, height_pdf - 160)

    # Guardar
    c.save()

    buffer.seek(0)
    pdf: bytes = buffer.getvalue()

    response.write(pdf)
    file_data = ContentFile(pdf)

    reportpdf = PDFFile()
    reportpdf.upload.save('REPORTE N°'+ str(id_rep)+'.pdf', file_data, save=False)
    reportpdf.save()

    pdfrelation = PDFReport(
        pdffile = reportpdf,
        report = reportehecho[0]
        )

    pdfrelation.save()

    #return response

#MODIFICAR REPORTE 
@login_required(login_url="login")
def modifiedreport(request, id):

    #Objetos a mostrar
    apis = API.objects.all()
    contracts = Contract.objects.all()
    specialties = Specialty.objects.all()
    references = Reference.objects.all()
    activities = Activity.objects.all()
    subactivities = SubActivity.objects.all()
    equipments = Equipment.objects.all()
    nonconformities = NonConformity.objects.all()
    followings = Following.objects.all()
    preconditions = Precondition.objects.all()
    measures = Measure.objects.all()
    historic = Historical.objects.filter(report=id)
    report = Report.objects.filter(id=id)[0]
    equipmentamount = EquipmentAmount.objects.filter(report=id)
    reportimage = ReportImage.objects.filter(report=id)
    historical_reference = HistoricalReference.objects.filter(report=id)

    id_contract = historic[0].activity.contract_id
    contract_report = Contract.objects.filter(id=id_contract)[0]
    
    i = 0
    specialty_id = historic[0].specialty_id
    id_refs = []
    id_hists = []

    for ref in historical_reference:
        i += 1
        id_refs.append({
                'id_ref_hist': ref.id,
                'id_referencia': ref.reference.id,
                'nombre_referencia': ref.reference.reference_name,
                'descripcion': ref.description,
                'id_table': "tr-ref" + str(i)
            })

    for hist in historic:

        id_hists.append({
                'id_html': 'hist-' + str(hist.id),
                'id': hist.id
            })

    assert isinstance(request, HttpRequest)
    return render(
        request,
        'form-report/modifiedreport.html',
        {
            'title':'Modificación de Informe Diario',
            'year': datetime.now().year,
            'apis': apis,
            'contracts': contracts,
            'specialties': specialties,
            'references': references,
            'activities': activities,
            'subactivities': subactivities,
            'equipments': equipments,
            'nonconformities': nonconformities,
            'followings': followings,
            'preconditions': preconditions,
            'measures': measures,
            'historical': historic,
            'report': report,
            'equipmentamount': equipmentamount,
            'reportimage' : reportimage,
            'historical_reference' : id_refs,
            'id_contract': id_contract,
            'id_historic': id_hists,
            'contract_report': contract_report,
            'specialty_id': specialty_id
    })
    
@csrf_exempt
@login_required(login_url="login")
def modifiedactualreport(request, *args, **kwargs):

    if request.method == "POST":

        action = request.POST.get('action')
        hd = request.POST.getlist('historico[]')
        rd = request.POST.getlist('referencias[]')
        ed = request.POST.getlist('equipos[]')
        rpd = request.POST.getlist('reporte[]')

        if action == 'save_report':

            data = {}

            historicaldata = json.loads(hd[0])
            referencedata = json.loads(rd[0])
            reportdata = json.loads(rpd[0])
            equipmentdata = json.loads(ed[0])

            rp_id = 0

            for i in reportdata:

                follow = Following.objects.get(pk = int(i['id_seguimiento']))

                old_report = Report.objects.get(pk = int(i['report_id']))
                    
                    
                if old_report.deviation_detected != i['desviacion']:

                    old_report.deviation_detected = i['desviacion']
                    old_report.save()
                
                if old_report.action_plan != i['plandeaccion']:

                    old_report.action_plan = i['plandeaccion']
                    old_report.save()

                if old_report.evidence_obs != i['evidencia_obs']:

                    old_report.evidence_obs = i['evidencia_obs']
                    old_report.save()

                if old_report.following_id != int(i['id_seguimiento']):
                    
                    old_report.following_id = int(i['id_seguimiento'])
                    old_report.save()

                id_report = int(i['report_id'])
                rp_id = int(i['report_id'])

            reporte = Report.objects.get(pk=int(i['report_id']))
                  
            for i in historicaldata:

                if int(i['id_historico']) == 0:

                    subact = SubActivity.objects.get(pk = int(i['id_subactividad']))

                    histo = Historical.objects.filter(subactivity = int(i['id_subactividad']))
                    diff = 0
                    for hs in histo:
                        diff += hs.real_amount

                    mea = Measure.objects.get(pk = int(i['id_medida']))
                    noncon = NonConformity.objects.get(pk = int(i['id_conformidad']))
                    spec = Specialty.objects.get(pk = int(i['id_especialidad']))
                    precon = Precondition.objects.get(pk = int(i['id_precondicion']))
                    actype = ActivityType.objects.get(pk = int(i['id_actividad_type']))
                    activi = Activity.objects.get(pk = int(i['id_actividad']))

                    historical = Historical(
                        real_amount = i['cantidad_real'],
                        subactivity_no_program = i['subactivity_no_program'],
                        subactivity = subact,
                        measure = mea,
                        nonconformity = noncon,
                        specialty= spec,
                        precondition= precon,
                        user= request.user,
                        activitytype= actype,
                        report = reporte,
                        no_program_total = i['total_estimado'],
                        no_program_refday = i['referencia_diaria'],
                        no_program_total_acu = i['total_acumulado'],
                        activity = activi,
                        difference = diff
                        )

                    historical.save()

                else:
                    #VERIFICAR SI EL HISTORICO FUE ELIMINADO AL MODIFICARLO 
                    old_historic = Historical.objects.get(pk = int(i['id_historico']))

                    if old_historic.real_amount != float(i['cantidad_real'].replace(",", ".")):
                        #REALIZAR DESCUENTO DE LA CANTIDAD PARA LA SUMA QUE SE HACE EN EL TOTAL ACUMULADO, 
                        #COMO TAMBIÉN PARA LOS HISTORICOS POSTERIORES A ESA FECHA
                        old_historic.real_amount = float(i['cantidad_real'].replace(",", "."))
                        old_historic.save()

                    if old_historic.no_program_total != float(i['total_estimado'].replace(",", ".")):

                        old_historic.no_program_total = float(i['total_estimado'].replace(",", "."))
                        old_historic.save()

                    if old_historic.no_program_refday != float(i['referencia_diaria'].replace(",", ".")):

                        old_historic.no_program_refday = float(i['referencia_diaria'].replace(",", "."))
                        old_historic.save()

                    if old_historic.no_program_total_acu != float(i['total_acumulado'].replace(",", ".")):

                        old_historic.no_program_total_acu = float(i['total_acumulado'].replace(",", "."))
                        old_historic.save()

                    if old_historic.subactivity_no_program != i['subactivity_no_program']:

                        old_historic.subactivity_no_program = i['subactivity_no_program']
                        old_historic.save()
                        
                    if old_historic.subactivity_id != int(i['id_subactividad']):

                        old_historic.subactivity_id = int(i['id_subactividad'])
                        old_historic.save()                    
                        
                    if old_historic.measure_id != int(i['id_medida']):

                        old_historic.measure_id = int(i['id_medida'])
                        old_historic.save()       
                        
                    if old_historic.nonconformity_id != int(i['id_conformidad']):

                        old_historic.nonconformity_id = int(i['id_conformidad'])
                        old_historic.save()     

                    if old_historic.specialty_id != int(i['id_especialidad']):

                        old_historic.specialty_id = int(i['id_especialidad'])
                        old_historic.save()             

                    if old_historic.precondition_id != int(i['id_precondicion']):

                        old_historic.precondition_id = int(i['id_precondicion'])
                        old_historic.save()               

                    if old_historic.activity_id != int(i['id_actividad']):

                        old_historic.activity_id = int(i['id_actividad'])
                        old_historic.save()

            if len(rd) > 0:

                for i in referencedata:

                    if int(i['id_ref_hist']) == 0:

                        ref = Reference.objects.get(pk = int(i['referencia_id']))

                        reference = HistoricalReference(
                            description = i['descripcion'],
                            report = reporte,
                            reference = ref
                            )

                        reference.save()

                    else:

                        old_ref = HistoricalReference.objects.get(pk = int(i['id_ref_hist']))

                        if old_ref.description != i['descripcion']:

                            old_ref.description = i['descripcion']
                            old_ref.save()

                        if old_ref.reference_id != int(i['referencia_id']):

                            old_ref.reference_id = int(i['referencia_id'])
                            old_ref.save()



            if len(ed) > 0:

                for i in equipmentdata:

                    if int(i['id_equipohist']) == 0:
                        act = Activity.objects.get(pk = int(i['id_actividad']))
                        equi = Equipment.objects.get(pk = int(i['id_equipo']))

                        equipment = EquipmentAmount(
                                equipment_amount = i['cantidad'],
                                direct_endowment = i['dot_directa'],
                                direct_reference = i['dot_referen'],
                                indirect_endowment = i['dot_indirecta'],
                                activity = act,
                                equipment = equi,
                                report = reporte
                            )
                        equipment.save()

                    else:

                        old_equi = EquipmentAmount.objects.get(pk=int(i['id_equipohist']))

                        if old_equi.equipment_amount != int(i['cantidad']):

                            old_equi.equipment_amount = int(i['cantidad'])
                            old_equi.save()     
                        
                        if old_equi.direct_endowment != int(i['dot_directa']):

                            old_equi.direct_endowment = int(i['dot_directa'])
                            old_equi.save()
                        
                        if old_equi.direct_reference != int(i['dot_referen']):

                            old_equi.direct_reference = int(i['dot_referen'])
                            old_equi.save()
                        
                        if old_equi.indirect_endowment != int(i['dot_indirecta']):

                            old_equi.indirect_endowment = int(i['dot_indirecta'])
                            old_equi.save()          

                        if old_equi.activity_id != int(i['id_actividad']):

                            old_equi.activity_id = int(i['id_actividad'])
                            old_equi.save()

                        if old_equi.equipment_id != int(i['id_equipo']):

                            old_equi.equipment_id = int(i['id_equipo'])
                            old_equi.save()

            if rp_id > 0:
                    
                data = {
                    'submitted': 1,
                    'id_report': rp_id
                    }

                if reportdata[0]['img_exist'] == 0:

                    create_pdf(rp_id)

            else:
                data = {
                    'submitted': 0
                    }

            return JsonResponse(data)

        
        if len(request.FILES.getlist('image')) > 0:

            image = request.FILES.getlist('image')
            observation = request.POST.getlist('observation')
            subactivity_image = request.POST.getlist('image_subactivity')
            id_img = request.POST.getlist('report')

            report_for_id = ReportImage.objects.get(id=int(id_img[0]))

            id_rep = report_for_id.report_id

            repor = Report.objects.get(pk = id_rep)
            obser = 0
            
            for i in image:

                if int(id_img[obser]) == 0:
                    subactividad_img = SubActivity.objects.get(id=int(subactivity_image[obser]))     
                
                    imagen = ImgReport(
                        image = i,
                        description = observation[obser]
                        )

                    imagen.save()

                    imagereport = ReportImage(
                        report = repor,
                        image = imagen,
                        subactivity = subactividad_img
                        )

                    imagereport.save()

                    obser = obser + 1

                else:
                    report_img = ReportImage.objects.get(id=int(id_img[obser]))

                    if report_img.subactivity_id != int(subactivity_image[obser]):

                        report_img.subactivity_id = int(subactivity_image[obser])
                        report_img.save()        
                        
                    if report_img.image.description != observation[obser]:

                        report_img.image.description = observation[obser]
                        report_img.save()

                    if report_img.image.image != i:
                        img_new = ImgReport(
                                image = i,
                                description = observation[obser]
                            )
                        img_new.save()

                        report_img.image_id = img_new.id
                        report_img.save()

                    obser = obser + 1


            if id_rep > 0:
                    
                data = {
                    'submitted': 1,
                    'id_report': id_rep
                    }

                create_pdf(id_rep)

            else:
                data = {
                    'submitted': 0
                    }

            return JsonResponse(data)

    else:
        return HttpResponse("<h2>No es posible agregar el reporte, verifique los datos correctamente</h2>")

